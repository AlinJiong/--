from numpy import require
import snscrape.modules.twitter as st
from sympy import im
import xlwt
import re
import io, urllib
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
import io, urllib
import requests

wb = Workbook()
ws = wb.active

ws.cell(1, 1, '时间')
ws.cell(1, 2, '文字内容及图片链接')
ws.cell(1, 3, '引用内容链接')
ws.cell(1, 4, '图片')
ws.cell(1, 5, '图片链接')

for i, tweet in enumerate(
        st.TwitterSearchScraper('from:optimismPBC').get_items()):
    tweet_list = [str(tweet.date), str(tweet.content), str(tweet.quotedTweet)]
    for j in range(2, 5):
        ws.cell(row=i + 2, column=j, value=tweet_list[j - 2])

    mediaUrl = tweet.media
    if mediaUrl == None:
        ws.cell(row=i + 2, column=4, value=' ')
        ws.cell(row=i + 2, column=5, value=' ')
    else:
        print(mediaUrl)

        try:
            imgUrl = re.findall('https://.*?=small', str(mediaUrl))[0]
        except:
            imgUrl = re.findall('https://.*?jpg', str(mediaUrl))[0]

        print(imgUrl)

        headers = {
            'User-Agent':
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/97.0.4692.71 Safari/537.36',
            'Content-Type': "application/x-www-form-urlencoded"
        }

        image = requests.get(url=imgUrl, headers=headers, timeout=10)
        image_data = io.BytesIO(image.content)

        img = Image(image_data)
        # imgsize = (img.width, img.height)
        imgsize = (720 / 4, 1280 / 4)

        img.width, img.height = imgsize

        ws.row_dimensions[i + 2].height = imgsize[1]  # 修改列第i+1行的高度

        ws.add_image(img, 'D' + str(i + 2))  # 图片 插入 A1 的位置上
        ws.cell(row=i + 2, column=5, value=str(mediaUrl))

wb.save('out.xlsx')  # 新的结果保存输出
wb.close()
